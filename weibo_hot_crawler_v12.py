# -*- coding: utf-8 -*-
"""
微博热搜榜 + 智搜(AI 总结)爬虫  v4.7
================================
v4.7 改动:按蓝色标签过滤广告位——检查每行是否有蓝色背景或
         蓝色小文字元素(商业推广位的蓝色"官宣/重磅/热推"等
         标签)。不再依赖关键词,避免真实热搜标题含"官宣"被误判。

v4.6 改动:思考去除规则改用"真·引用"识别——
  真答案的 @ 永远在行末作为来源标注(如"...@新华社"),
  而思考里的 @ 是正文中的 @某人称呼(如"@某人 指出...")。
  规则:找第一个满足以下之一的行,从那里开始保留:
    (a) 行末是 @xxx 引用标注(@后最多15字符+省略号/句号/空格)
    (b) 严格章节标题(一、/1️⃣/emoji)
  两个信号取更早的那个;若命中的是标题,再尝试往前退一行找总起句
  (退到的行必须也是"真·引用"才保留)。

v4 改动(相对 v3):
  - 智搜 JS 提取时,先在 DOM 层删除 img/video/figure 等媒体元素
    及 class 含 video/card/thumb/cover/pic/image 的容器,避免视频卡片
    里的用户名标签污染文本
  - 文本层智能清洗:
      * 删除时间戳行
      * 删除 @用户名 后面紧跟的互动数字串(如 @第一现场67815 → @第一现场)
      * 删除行尾纯数字
      * 智能识别孤立用户名:短行只在"前后都是正文"时才视为小标题保留,
        否则按孤立用户名剔除——这样既能保留 "事件核心矛盾" 等小标题,
        又能干掉视频缩略图带出来的用户名行

依赖:
  pip install playwright openpyxl
"""

import json
import random
import re
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ============================================================
# 配置区
# ============================================================
USER_DATA_DIR = "./wb_user_data"
OUTPUT_DIR = Path("./output")
DEBUG_DIR = Path("./debug")
MAX_ITEMS = 10
START_URL = "https://s.weibo.com/top/summary"
LOGIN_WAIT_SECONDS = 180

INTERVAL_MIN = 4.0
INTERVAL_MAX = 8.0
REST_EVERY = 8
REST_SECONDS = 30
COOLDOWN_ON_418 = 60

ZHISOU_RENDER_TIMEOUT = 25000
ZHISOU_EXTRA_WAIT = 2.5

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0")

NAV_NOISE = {
    "综合", "智搜", "实时", "热门", "视频", "图片", "高级搜索",
    "搜索结果", "微博热搜", "我的", "热搜", "刷新",
    "查看完整热搜榜单", "分享", "我有补充", "反馈", "复制",
    "TOP", "返回顶部", "回答·深度思考",
}


# ============================================================
# 通用工具
# ============================================================
def human_sleep(a=INTERVAL_MIN, b=INTERVAL_MAX):
    time.sleep(random.uniform(a, b))


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def safe_text(locator):
    try:
        return (locator.inner_text(timeout=2000) or "").strip()
    except Exception:
        return ""


def parse_read_count(text):
    m = re.search(r"阅读[量数]?\s*([\d\.]+\s*[万亿]?)", text)
    return m.group(1).strip() if m else ""


def is_blocked_page(page):
    try:
        body = safe_text(page.locator("body"))
        if "无法正常运作" in body or "ERROR 418" in body or "HTTP ERROR" in body:
            return True
    except Exception:
        pass
    return False


def dump_debug(page, name):
    DEBUG_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%H%M%S")
    try:
        (DEBUG_DIR / f"{name}_{ts}.html").write_text(page.content(), encoding="utf-8")
        page.screenshot(path=str(DEBUG_DIR / f"{name}_{ts}.png"), full_page=True)
    except Exception:
        pass


# ============================================================
# 登录 / 列表
# ============================================================
LOGIN_MARKER = "[class*='gn_name'], [class*='Nav_name'], img[class*='avatar']"


def ensure_logged_in(page):
    page.goto(START_URL, wait_until="domcontentloaded")
    time.sleep(2)
    try:
        if page.locator(LOGIN_MARKER).first.is_visible(timeout=2000):
            log("✓ 检测到已登录态,直接开始")
            return
    except Exception:
        pass
    log(f"⚠ 未登录。请扫码登录,最多等 {LOGIN_WAIT_SECONDS} 秒……")
    deadline = time.time() + LOGIN_WAIT_SECONDS
    while time.time() < deadline:
        try:
            if page.locator(LOGIN_MARKER).first.is_visible(timeout=1000):
                log("✓ 登录成功")
                time.sleep(2)
                return
        except Exception:
            pass
        time.sleep(2)
    raise RuntimeError("登录超时")


def collect_hot_list(page):
    page.goto(START_URL, wait_until="domcontentloaded")
    page.wait_for_selector("table tbody tr", timeout=15000)
    time.sleep(1.5)
    rows = page.locator("table tbody tr:not(.hidden_topic)")
    total = rows.count()
    log(f"共发现 {total} 条榜单行(含可能的广告)")

    # 判定广告的 JS:只检查最右边的标签列(td.td-03 或最后一个 td)
    # 里是否有蓝色元素。标题列本身是蓝色链接文字,不能整行扫。
    IS_AD_JS = r"""
    (row) => {
        const isBlue = (rgbStr) => {
            const m = rgbStr && rgbStr.match(/rgba?\((\d+),\s*(\d+),\s*(\d+)(?:,\s*([\d.]+))?/);
            if (!m) return false;
            const r = +m[1], g = +m[2], b = +m[3];
            const a = m[4] === undefined ? 1 : +m[4];
            if (a < 0.3) return false;
            return b > 100 && b > r + 30 && b > g + 30;
        };
        // 优先找 td.td-03,找不到就取最后一个 td
        let cell = row.querySelector('td.td-03');
        if (!cell) {
            const tds = row.querySelectorAll('td');
            cell = tds[tds.length - 1];
        }
        if (!cell) return false;
        // 检查该单元格自身和所有子元素
        const all = [cell, ...cell.querySelectorAll('*')];
        for (const el of all) {
            const s = window.getComputedStyle(el);
            if (isBlue(s.backgroundColor)) return true;
            if (isBlue(s.color)) {
                const txt = (el.innerText || '').trim();
                if (txt && txt.length <= 6) return true;
            }
        }
        return false;
    }
    """

    items = []
    skipped_ads = 0
    for i in range(total):
        if len(items) >= MAX_ITEMS:
            break
        row = rows.nth(i)

        title = safe_text(row.locator("td.td-02 a"))

        # 按颜色判定广告
        try:
            is_ad = row.evaluate(IS_AD_JS)
        except Exception:
            is_ad = False
        if is_ad:
            skipped_ads += 1
            log(f"  跳过广告: {title}")
            continue

        rank = safe_text(row.locator("td.td-01"))
        heat = safe_text(row.locator("td.td-02 span"))
        href = ""
        try:
            href = row.locator("td.td-02 a").first.get_attribute("href") or ""
            if href.startswith("/"):
                href = "https://s.weibo.com" + href
        except Exception:
            pass
        if title:
            items.append({"rank": rank or str(len(items) + 1), "title": title,
                          "heat": heat, "url": href})

    log(f"过滤广告 {skipped_ads} 条,将处理 {len(items)} 条真实热搜")
    return items


# ============================================================
# 详情页阅读量
# ============================================================
def goto_with_retry(page, url):
    for attempt in (1, 2):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=20000)
        except PWTimeout:
            return "error:timeout"
        except Exception as e:
            return f"error:{e}"
        time.sleep(2)
        if is_blocked_page(page):
            if attempt == 1:
                log(f"  ⚠ 触发 418 风控,冷却 {COOLDOWN_ON_418} 秒后重试……")
                time.sleep(COOLDOWN_ON_418)
                continue
            return "blocked"
        return "ok"
    return "blocked"


def get_read_count(page, item):
    if not item["url"] or item["url"].startswith("javascript:"):
        return ""
    if goto_with_retry(page, item["url"]) != "ok":
        return ""
    time.sleep(1.5)
    return parse_read_count(safe_text(page.locator("body")))


# ============================================================
# 智搜:JS 提取(含 DOM 层媒体清理)
# ============================================================
JS_EXTRACT = r"""
() => {
    // 1. 找到 "内容由AI生成" 文本节点
    const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
    let markerEl = null;
    let n;
    while (n = walker.nextNode()) {
        const t = (n.textContent || '').trim();
        if (t === '内容由AI生成' || t.includes('内容由AI生成')) {
            markerEl = n.parentElement;
            break;
        }
    }
    if (!markerEl) return null;

    // 2. 向上回溯找正文容器
    let container = markerEl;
    let bestEl = null;
    for (let i = 0; i < 20; i++) {
        if (!container.parentElement) break;
        container = container.parentElement;
        const text = container.innerText || '';
        if (text.length > 800 && (text.includes('总结') || text.includes('一、') || text.includes('核心') || text.includes('事件'))) {
            bestEl = container;
            if (text.length > 8000) break;
        }
    }
    if (!bestEl) return null;

    // 3. 克隆后在副本里删除媒体元素和视频卡片容器
    const clone = bestEl.cloneNode(true);
    const tagsToRemove = ['img', 'video', 'picture', 'canvas', 'svg', 'figure', 'iframe'];
    tagsToRemove.forEach(tag => {
        clone.querySelectorAll(tag).forEach(e => e.remove());
    });
    const classKeywords = ['video', 'Video', 'card', 'Card', 'thumb', 'Thumb',
                           'cover', 'Cover', 'pic', 'Pic', 'image', 'Image',
                           'media', 'Media', 'avatar', 'Avatar'];
    classKeywords.forEach(kw => {
        try {
            clone.querySelectorAll('[class*="' + kw + '"]').forEach(e => e.remove());
        } catch (e) {}
    });

    return clone.innerText;
}
"""


def open_zhisou(page, item):
    """先尝试点侧栏的『智搜』,失败则直接构造 aisearch URL 跳转"""
    try:
        btn = page.locator("text=智搜").first
        if btn.count() > 0:
            btn.click(timeout=5000)
            time.sleep(2)
            if "aisearch" in page.url:
                return True
    except Exception:
        pass
    try:
        q = quote(f"#{item['title']}#")
        ai_url = f"https://s.weibo.com/aisearch?q={q}&Refer=weibo_aisearch&t=31"
        page.goto(ai_url, wait_until="domcontentloaded", timeout=20000)
        time.sleep(2)
        return True
    except Exception as e:
        log(f"  打开智搜失败: {e}")
        return False


# ============================================================
# 智搜:文本清洗
# ============================================================
def is_content_line(s):
    """判断一行是不是『正文』(长行 / 带句末标点 / 编号小标题)"""
    if not s:
        return False
    if len(s) > 25:
        return True
    if re.search(r"[。:!?,;:!?,;]", s):
        return True
    if re.match(r"^[一二三四五六七八九十百]+[、.]", s):
        return True
    if re.match(r"^\d+[\.\)、️⃣]", s):
        return True
    if re.match(r"^[●•·▪◆■]", s):
        return True
    return False


def clean_zhisou_text(raw):
    """整体清洗:去时间戳/数字徽章/孤立用户名簇,保留正文和小标题"""
    if not raw:
        return ""

    # ---- 1. 截尾:在这些标记之前停止 ----
    end_idx = len(raw)
    for marker in ["你可能想问", "信源追溯", "内容由AI生成"]:
        i = raw.find(marker)
        if i > 0:
            end_idx = min(end_idx, i)
    text = raw[:end_idx]

    # ---- 2. 截头:跳过 "回答·深度思考" 那一行 ----
    head_idx = 0
    for marker in ["查看最新", "深度思考"]:
        i = text.find(marker)
        if i >= 0:
            nl = text.find("\n", i)
            if nl > head_idx:
                head_idx = nl + 1
    text = text[head_idx:]

    # ---- 2.5 删除 AI 思考段落 ----
    # 真答案的开头必然是下面两种之一:
    #   (a) 一行总起句,包含 @来源
    #   (b) 严格章节标题行(一、/1️⃣/emoji 开头)
    # 而思考段落既不引用 @来源,也不用 "一、" 这种严格标题。
    # 所以:扫一遍,找第一个满足 (a) 或 (b) 的行,取最早的,
    # 从它开始保留。若该行是标题,再尝试往前退一行——只有退到
    # 的那行含 @ 才保留,否则直接从标题开始。
    _strict_heading = re.compile(
        r"^("
        r"[一二三四五六七八九十]、"
        r"|\d\uFE0F\u20E3"
        r"|[\U0001F300-\U0001F9FF\u2600-\u27BF\uFE0F]+\s*\S"
        r")"
    )
    # "真·引用 @ 标注"的判定:行尾有 @xxx (最多15个非 @ 字符,可带省略号)
    # 这样能区分"引用来源"(答案特征)和"正文 @某人 指出"(思考特征)
    _at_citation = re.compile(r"@[^@\s]{1,15}[\.。…\s]*$")
    _lines_tmp = text.split("\n")
    _at_idx = -1
    _head_idx = -1
    for _i, _ln in enumerate(_lines_tmp):
        _s = _ln.strip()
        if _at_idx < 0 and _at_citation.search(_s):
            _at_idx = _i
        if _head_idx < 0 and _strict_heading.match(_s):
            _head_idx = _i
        if _at_idx >= 0 and _head_idx >= 0:
            break
    # 选更早的那个
    _candidates = [x for x in (_at_idx, _head_idx) if x >= 0]
    if _candidates:
        _pick = min(_candidates)
        # 若选中的是章节标题,往前退一行找总起句——但只有那行是"真·引用"才退
        if _pick == _head_idx and _pick != _at_idx:
            for _j in range(_pick - 1, -1, -1):
                if _lines_tmp[_j].strip():
                    if _at_citation.search(_lines_tmp[_j]):
                        _pick = _j
                    break
        text = "\n".join(_lines_tmp[_pick:])

    # ---- 3. 逐行第一轮清洗 ----
    # AI 思考段落:以 "嗯,"(嗯+逗号/顿号/空格)开头,可能跨多行。
    # 持续丢弃,直到遇到一行明确属于答案的内容——判定信号:
    #   (a) 行内包含 "@用户名"(AI 答案几乎都会标注来源,思考不会),或
    #   (b) 以章节标题开头(一、/二、/1️⃣/emoji)
    _ans_heading = re.compile(
        r"^("
        r"[一二三四五六七八九十]、"
        r"|\d\uFE0F\u20E3"
        r"|[\U0001F300-\U0001F9FF\u2600-\u27BF\uFE0F]+\s*\S"
        r")"
    )
    def _is_answer_line(s):
        if "@" in s:
            return True
        if _ans_heading.match(s):
            return True
        return False

    lines = []
    in_thinking = False
    for line in text.split("\n"):
        s = line.strip()
        if not in_thinking and s.startswith("嗯") and len(s) > 2 and s[1] in "，,、 ":
            in_thinking = True
            continue
        if in_thinking:
            if _is_answer_line(s):
                in_thinking = False  # 进入答案,本行保留,继续走下面
            else:
                continue              # 还在思考段落,丢
        if not s:
            continue
        # 侧栏导航词
        if s in NAV_NOISE:
            continue
        # 时间戳
        if re.match(r"^时间[::]", s):
            continue
        if re.match(r"^\d+[分小时天月]\w*前$", s):
            continue
        # 纯数字行(互动徽章被换行成自己的一行)
        if re.fullmatch(r"\d{1,6}", s):
            continue
        # 行内: @用户名 后面紧跟的数字串删掉
        # 例如 "@第一现场67815" -> "@第一现场"
        s = re.sub(r"(@[^\s@\d,。、:;!?,。、:;!?]+?)(\d{1,15})(?=\s|@|$|[,。、:;!?,。、:;!?])", r"\1", s)
        # 行尾的数字串
        s = re.sub(r"\s*\d{1,15}\s*$", "", s)
        # 多个空白合并
        s = re.sub(r"\s{2,}", " ", s).strip()
        if s:
            lines.append(s)

    # ---- 4. 智能识别孤立用户名 ----
    # 短行(非正文)只有在前后两侧都是正文时才视为小标题保留,
    # 否则视为视频缩略图带出来的用户名,丢弃
    final = []
    n = len(lines)
    for i, s in enumerate(lines):
        if is_content_line(s):
            final.append(s)
            continue
        prev_ok = (i == 0) or is_content_line(lines[i - 1])
        next_ok = (i == n - 1) or is_content_line(lines[i + 1])
        if prev_ok and next_ok:
            final.append(s)
        # else: drop

    return "\n".join(final).strip()


# ============================================================
# 智搜主流程
# ============================================================
def crawl_zhisou(page, item):
    if not open_zhisou(page, item):
        return None, "未能打开智搜页"

    # 等 AI 流式输出完成
    try:
        page.wait_for_function(
            """() => {
                const t = document.body.innerText || '';
                return t.includes('内容由AI生成') || t.includes('你可能想问');
            }""",
            timeout=ZHISOU_RENDER_TIMEOUT,
        )
    except PWTimeout:
        log("  ⚠ 等待 AI 内容超时,尝试用现有内容提取")
    time.sleep(ZHISOU_EXTRA_WAIT)

    # 策略 1:JS 定位 + DOM 媒体清理
    raw = None
    try:
        raw = page.evaluate(JS_EXTRACT)
    except Exception as e:
        log(f"  JS 提取异常: {e}")

    # 策略 2:兜底 body 全文
    if not raw:
        try:
            raw = page.locator("body").inner_text(timeout=5000)
        except Exception:
            return None, "页面文本读取失败"

    text = clean_zhisou_text(raw)

    if not text or len(text) < 30:
        dump_debug(page, f"zhisou_short_{item['rank']}")
        return None, f"提取内容过短({len(text)} 字),已 dump 调试文件"

    return text, "ok"


# ============================================================
# 输出
# ============================================================
def save_json(records, path):
    path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def save_excel(records, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "微博热搜"
    headers = ["排名", "标题", "热度", "阅读量",
               "智搜AI总结", "智搜状态", "详情链接", "抓取时间"]
    ws.append(headers)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4472C4")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in records:
        ws.append([
            r.get("rank", ""),
            r.get("title", ""),
            r.get("heat", ""),
            r.get("read_count", ""),
            r.get("ai_summary") or "",
            r.get("ai_summary_status", ""),
            r.get("url", ""),
            r.get("crawled_at", ""),
        ])

    widths = [6, 38, 12, 12, 90, 22, 50, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def save_all(records, json_path, xlsx_path):
    try:
        save_json(records, json_path)
        save_excel(records, xlsx_path)
    except Exception as e:
        log(f"  ⚠ 保存中间结果失败(不影响继续): {e}")


# ============================================================
# 主程序
# ============================================================
def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    DEBUG_DIR.mkdir(exist_ok=True)

    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    json_path = OUTPUT_DIR / f"weibo_hot_{stamp}.json"
    xlsx_path = OUTPUT_DIR / f"weibo_hot_{stamp}.xlsx"
    log(f"本次输出文件: {json_path.name} / {xlsx_path.name}")

    results = []
    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(
            user_data_dir=USER_DATA_DIR,
            headless=False,
            channel="msedge",
            user_agent=UA,
            viewport={"width": 1400, "height": 900},
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = ctx.new_page()

        try:
            ensure_logged_in(page)
            hot_list = collect_hot_list(page)

            for idx, item in enumerate(hot_list, 1):
                log(f"[{idx}/{len(hot_list)}] {item['title']}")
                record = {**item,
                          "read_count": "",
                          "ai_summary": None,
                          "ai_summary_status": "",
                          "crawled_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
                try:
                    record["read_count"] = get_read_count(page, item)
                    summary, status = crawl_zhisou(page, item)
                    record["ai_summary"] = summary
                    record["ai_summary_status"] = status
                    if summary:
                        log(f"  ✓ 智搜 {len(summary)} 字")
                    else:
                        log(f"  ✗ 智搜失败: {status}")
                except Exception as e:
                    log(f"  抓取异常: {e}")
                    record["ai_summary_status"] = f"异常: {e}"

                results.append(record)
                save_all(results, json_path, xlsx_path)

                if idx % REST_EVERY == 0 and idx < len(hot_list):
                    log(f"  💤 已爬 {idx} 条,休息 {REST_SECONDS} 秒……")
                    time.sleep(REST_SECONDS)
                else:
                    human_sleep()

            ok = sum(1 for r in results if r.get("ai_summary"))
            log(f"全部完成。共 {len(results)} 条,其中 {ok} 条成功获取智搜")

        except KeyboardInterrupt:
            log("⛔ 收到 Ctrl+C,中止抓取。已抓到的数据已保存。")
        except Exception as e:
            log(f"⛔ 主流程异常: {e}。已抓到的数据已保存。")
        finally:
            if results:
                save_all(results, json_path, xlsx_path)
                log(f"✓ 最终保存:共 {len(results)} 条 → {xlsx_path.name}")
            try:
                ctx.close()
            except Exception:
                pass


if __name__ == "__main__":
    main()
